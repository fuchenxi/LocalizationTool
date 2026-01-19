#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
长度对比标签页
对比不同语言的 value 长度，找出变长的字段
设计风格：遵循 Apple Human Interface Guidelines
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
    QComboBox, QPushButton, QTextEdit,
    QTableWidget, QTableWidgetItem,
    QHeaderView, QCheckBox, QDoubleSpinBox,
    QDialog, QDialogButtonBox, QScrollArea,
    QFrame, QApplication, QFileDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QBrush, QFont
from utils.theme import get_theme_colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class LanguageSelectorDialog(QDialog):
    """语言选择对话框"""
    
    def __init__(self, languages: list, selected: list, parent=None):
        super().__init__(parent)
        self.languages = languages
        self.selected = selected.copy()
        self.checkboxes = {}
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("选择目标语言")
        self.setMinimumWidth(280)
        self.setStyleSheet("""
            QDialog {
                background: white;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)
        
        # 提示
        hint = QLabel("选择要检测长度变化的语言：")
        hint.setStyleSheet("font-size: 13px; color: #666; margin-bottom: 8px;")
        layout.addWidget(hint)
        
        # 滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(300)
        scroll.setStyleSheet("border: none;")
        
        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.setSpacing(8)
        
        for lang in self.languages:
            cb = QCheckBox(lang)
            cb.setChecked(lang in self.selected)
            cb.setStyleSheet("""
                QCheckBox {
                    font-size: 14px;
                    padding: 6px 0;
                }
                QCheckBox::indicator {
                    width: 20px;
                    height: 20px;
                    border-radius: 4px;
                    border: 1.5px solid #C7C7CC;
                }
                QCheckBox::indicator:checked {
                    background: #007AFF;
                    border: none;
                    image: url(data:image/svg+xml;base64,PHN2ZyB3aWR0aD0iMTIiIGhlaWdodD0iMTIiIHZpZXdCb3g9IjAgMCAxMiAxMiIgZmlsbD0ibm9uZSIgeG1sbnM9Imh0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnIj4KPHBhdGggZD0iTTEwIDNMNC41IDguNUwyIDYiIHN0cm9rZT0id2hpdGUiIHN0cm9rZS13aWR0aD0iMiIgc3Ryb2tlLWxpbmVjYXA9InJvdW5kIiBzdHJva2UtbGluZWpvaW49InJvdW5kIi8+Cjwvc3ZnPgo=);
                }
            """)
            self.checkboxes[lang] = cb
            container_layout.addWidget(cb)
        
        container_layout.addStretch()
        scroll.setWidget(container)
        layout.addWidget(scroll)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)
        
        select_all_btn = QPushButton("全选")
        select_all_btn.clicked.connect(self.select_all)
        select_all_btn.setStyleSheet("""
            QPushButton {
                background: transparent;
                color: #007AFF;
                border: none;
                font-size: 13px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                color: #0051D5;
            }
        """)
        btn_layout.addWidget(select_all_btn)
        
        clear_btn = QPushButton("清除")
        clear_btn.clicked.connect(self.clear_all)
        clear_btn.setStyleSheet(select_all_btn.styleSheet())
        btn_layout.addWidget(clear_btn)
        
        btn_layout.addStretch()
        
        confirm_btn = QPushButton("确定")
        confirm_btn.clicked.connect(self.accept)
        confirm_btn.setStyleSheet("""
            QPushButton {
                background: #007AFF;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                font-weight: 500;
                padding: 10px 24px;
            }
            QPushButton:hover {
                background: #0051D5;
            }
        """)
        btn_layout.addWidget(confirm_btn)
        
        layout.addLayout(btn_layout)
    
    def select_all(self):
        for cb in self.checkboxes.values():
            cb.setChecked(True)
    
    def clear_all(self):
        for cb in self.checkboxes.values():
            cb.setChecked(False)
    
    def get_selected(self) -> list:
        return [lang for lang, cb in self.checkboxes.items() if cb.isChecked()]


class LengthCompareTab(QWidget):
    """长度对比标签页 - 现代简洁设计"""
    
    def __init__(self):
        super().__init__()
        self.colors = get_theme_colors()
        self.languages = []
        self.selected_languages = []
        self.results = {}
        self.sorted_results = []
        self.init_ui()
    
    def init_ui(self):
        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(24, 20, 24, 20)
        main_layout.setSpacing(16)
        
        # ========== 顶部配置区 ==========
        config_widget = QWidget()
        config_widget.setStyleSheet("""
            QWidget {
                background: #F5F5F7;
                border-radius: 12px;
            }
        """)
        config_layout = QHBoxLayout(config_widget)
        config_layout.setContentsMargins(16, 14, 16, 14)
        config_layout.setSpacing(16)
        
        # 目标语言选择
        lang_label = QLabel("目标语言")
        lang_label.setStyleSheet("font-size: 13px; color: #666; background: transparent;")
        config_layout.addWidget(lang_label)
        
        self.lang_selector_btn = QPushButton("选择语言...")
        self.lang_selector_btn.setMinimumWidth(140)
        self.lang_selector_btn.clicked.connect(self.show_language_selector)
        self.lang_selector_btn.setStyleSheet("""
            QPushButton {
                background: white;
                border: none;
                border-radius: 8px;
                padding: 8px 14px;
                font-size: 13px;
                color: #1D1D1F;
                text-align: left;
            }
            QPushButton:hover {
                background: #FAFAFA;
            }
        """)
        config_layout.addWidget(self.lang_selector_btn)
        
        # 分隔线
        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.VLine)
        sep1.setStyleSheet("background: #E5E5EA; max-width: 1px;")
        config_layout.addWidget(sep1)
        
        # 对比基准
        mode_label = QLabel("对比基准")
        mode_label.setStyleSheet("font-size: 13px; color: #666; background: transparent;")
        config_layout.addWidget(mode_label)
        
        self.compare_mode_combo = QComboBox()
        self.compare_mode_combo.addItems([
            "平均长度",
            "最大长度",
            "指定语言"
        ])
        self.compare_mode_combo.setMinimumWidth(120)
        self.compare_mode_combo.currentIndexChanged.connect(self.on_compare_mode_changed)
        self.compare_mode_combo.setStyleSheet("""
            QComboBox {
                background: white;
                border: none;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 13px;
                color: #1D1D1F;
            }
            QComboBox:hover {
                background: #FAFAFA;
            }
            QComboBox::drop-down {
                border: none;
                width: 24px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 5px solid #8E8E93;
                margin-right: 8px;
            }
            QComboBox QAbstractItemView {
                background: white;
                border: 1px solid #E5E5EA;
                border-radius: 8px;
                padding: 4px;
                selection-background-color: #007AFF;
            }
        """)
        config_layout.addWidget(self.compare_mode_combo)
        
        # 基准语言（默认隐藏）
        self.base_lang_combo = QComboBox()
        self.base_lang_combo.setMinimumWidth(100)
        self.base_lang_combo.setVisible(False)
        self.base_lang_combo.setStyleSheet(self.compare_mode_combo.styleSheet())
        config_layout.addWidget(self.base_lang_combo)
        
        # 分隔线
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.Shape.VLine)
        sep2.setStyleSheet("background: #E5E5EA; max-width: 1px;")
        config_layout.addWidget(sep2)
        
        # 阈值
        threshold_label = QLabel("阈值")
        threshold_label.setStyleSheet("font-size: 13px; color: #666; background: transparent;")
        config_layout.addWidget(threshold_label)
        
        self.min_diff_spinbox = QDoubleSpinBox()
        self.min_diff_spinbox.setMinimum(0.0)
        self.min_diff_spinbox.setMaximum(1000.0)
        self.min_diff_spinbox.setValue(0.0)
        self.min_diff_spinbox.setSuffix("%")
        self.min_diff_spinbox.setDecimals(0)
        self.min_diff_spinbox.setFixedWidth(80)
        self.min_diff_spinbox.setStyleSheet("""
            QDoubleSpinBox {
                background: white;
                border: none;
                border-radius: 8px;
                padding: 8px 10px;
                font-size: 13px;
                color: #1D1D1F;
            }
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {
                width: 0;
                border: none;
            }
        """)
        config_layout.addWidget(self.min_diff_spinbox)
        
        config_layout.addStretch()
        
        # 开始对比按钮
        self.compare_btn = QPushButton("开始对比")
        self.compare_btn.setEnabled(False)
        self.compare_btn.setStyleSheet("""
            QPushButton {
                background: #007AFF;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton:hover {
                background: #0051D5;
            }
            QPushButton:disabled {
                background: #C7C7CC;
            }
        """)
        config_layout.addWidget(self.compare_btn)
        
        main_layout.addWidget(config_widget)
        
        # ========== 统计栏 ==========
        stats_widget = QWidget()
        stats_layout = QHBoxLayout(stats_widget)
        stats_layout.setContentsMargins(4, 0, 4, 0)
        stats_layout.setSpacing(12)
        
        self.stats_label = QLabel("选择目标语言后开始对比")
        self.stats_label.setStyleSheet("font-size: 14px; color: #8E8E93;")
        stats_layout.addWidget(self.stats_label)
        
        stats_layout.addStretch()
        
        # 修改按钮文字和样式（约第 330 行）
        self.copy_btn = QPushButton("导出 Excel")
        self.copy_btn.setVisible(False)
        self.copy_btn.clicked.connect(self.export_to_excel)  # 改为 export_to_excel
        self.copy_btn.setStyleSheet("""
            QPushButton {
                background: transparent;
                color: #007AFF;
                border: none;
                font-size: 13px;
                padding: 6px 12px;
            }
            QPushButton:hover {
                color: #0051D5;
                background: #F5F5F7;
                border-radius: 6px;
            }
        """)
        stats_layout.addWidget(self.copy_btn)
        
        main_layout.addWidget(stats_widget)
        
        # ========== 结果区域 ==========
        self.result_container = QWidget()
        result_layout = QVBoxLayout(self.result_container)
        result_layout.setContentsMargins(0, 0, 0, 0)
        result_layout.setSpacing(0)
        
        # 空状态提示
        self.empty_widget = QWidget()
        empty_layout = QVBoxLayout(self.empty_widget)
        empty_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        self.empty_label = QLabel("选择目标语言，点击「开始对比」")
        self.empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.empty_label.setStyleSheet("""
            font-size: 15px; 
            color: #C7C7CC; 
            padding: 60px;
        """)
        empty_layout.addWidget(self.empty_label)
        result_layout.addWidget(self.empty_widget)
        
        # 结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(7)
        self.result_table.setHorizontalHeaderLabels([
            "Key", "英文 Value", "Value", "语言", "长度", "基准", "差异"
        ])
        self.result_table.setAlternatingRowColors(False)
        self.result_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.result_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.result_table.setShowGrid(False)
        self.result_table.verticalHeader().setVisible(False)
        self.result_table.setStyleSheet("""
            QTableWidget {
                border: none;
                background: transparent;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px 8px;
                border-bottom: 1px solid #F0F0F0;
            }
            QTableWidget::item:selected {
                background: #E8F0FE;
                color: #1D1D1F;
            }
            QTableWidget::item:hover {
                background: #FAFAFA;
            }
            QHeaderView::section {
                background: transparent;
                border: none;
                border-bottom: 1px solid #E5E5EA;
                padding: 10px 8px;
                font-size: 12px;
                font-weight: 600;
                color: #8E8E93;
                text-transform: uppercase;
            }
        """)
        
        # 设置列宽
        header = self.result_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # Key
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # 英文 Value
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # Value
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)    # 语言
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)    # 长度
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)    # 基准
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)    # 差异
        
        self.result_table.setColumnWidth(0, 150)  # Key
        self.result_table.setColumnWidth(3, 80)   # 语言
        self.result_table.setColumnWidth(4, 70)   # 长度
        self.result_table.setColumnWidth(5, 70)   # 基准
        self.result_table.setColumnWidth(6, 90)   # 差异
        
        self.result_table.setVisible(False)
        result_layout.addWidget(self.result_table)
        
        main_layout.addWidget(self.result_container, 1)
        
        # 隐藏的日志（保留接口兼容）
        self.compare_log_text = QTextEdit()
        self.compare_log_text.setVisible(False)
    
    def show_language_selector(self):
        """显示语言选择对话框"""
        dialog = LanguageSelectorDialog(self.languages, self.selected_languages, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.selected_languages = dialog.get_selected()
            self.update_lang_button_text()
    
    def update_lang_button_text(self):
        """更新语言按钮文字"""
        if not self.selected_languages:
            self.lang_selector_btn.setText("选择语言...")
        elif len(self.selected_languages) == 1:
            self.lang_selector_btn.setText(self.selected_languages[0])
        elif len(self.selected_languages) <= 3:
            self.lang_selector_btn.setText(", ".join(self.selected_languages))
        else:
            self.lang_selector_btn.setText(f"已选 {len(self.selected_languages)} 个语言")
    
    def on_compare_mode_changed(self, index: int):
        """对比模式改变"""
        self.base_lang_combo.setVisible(index == 2)
    
    def update_languages(self, languages: list):
        """更新语言列表"""
        self.languages = languages
        self.selected_languages = []
        self.update_lang_button_text()
        
        # 更新基准语言下拉框
        self.base_lang_combo.clear()
        self.base_lang_combo.addItems(languages)
        
        # 默认选择 en
        if 'en' in languages:
            self.base_lang_combo.setCurrentIndex(languages.index('en'))
    
    def get_selected_target_languages(self) -> list:
        """获取选中的目标语言"""
        return self.selected_languages
    
    def get_compare_mode(self) -> str:
        """获取对比模式"""
        index = self.compare_mode_combo.currentIndex()
        if index == 0:
            return "average"
        elif index == 1:
            return "max"
        elif index == 2:
            return "base_lang"
        return "average"
    
    def get_base_lang(self) -> str:
        """获取基准语言"""
        return self.base_lang_combo.currentText()
    
    def get_min_diff_percent(self) -> float:
        """获取最小差异百分比"""
        return self.min_diff_spinbox.value()
    
    def export_to_excel(self):
        """导出结果到 Excel"""
        if not self.sorted_results:
            return
        
        # 让用户选择保存位置
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出 Excel",
            "长度对比结果.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            # 创建 Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "长度对比结果"
            
            # 设置表头
            headers = ["Key", "英文 Value", "Value", "语言", "长度", "基准", "差异%"]
            ws.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color="F5F5F7", end_color="F5F5F7", fill_type="solid")
            header_font = Font(bold=True, size=12, color="1D1D1F")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 填充数据
            for _, data in self.sorted_results:
                base_len = data['base_length']
                if isinstance(base_len, float) and base_len.is_integer():
                    base_str = int(base_len)
                else:
                    base_str = round(base_len, 1)
                
                diff_percent = data['diff_percent']
                
                # 获取英文 Value
                en_value = ""
                if 'all_values' in data:
                    # 优先查找 'en'，如果没有则查找以 'en' 开头的语言代码
                    if 'en' in data['all_values']:
                        en_value = data['all_values']['en']['value']
                    else:
                        for lang_code in data['all_values'].keys():
                            if lang_code.startswith('en'):
                                en_value = data['all_values'][lang_code]['value']
                                break
                
                row = [
                    data['key'],
                    en_value,
                    data['target_value'],
                    data['target_lang'],
                    data['target_length'],
                    base_str,
                    f"+{diff_percent:.1f}%"
                ]
                ws.append(row)
                
                # 设置差异列的颜色（根据严重程度）
                diff_cell = ws.cell(row=ws.max_row, column=7)
                if diff_percent >= 100:
                    diff_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    diff_cell.font = Font(bold=True, color="FF3B30")
                elif diff_percent >= 50:
                    diff_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                    diff_cell.font = Font(bold=True, color="FF6B35")
                else:
                    diff_cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                    diff_cell.font = Font(bold=True, color="FF9500")
                
                diff_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 设置列宽
            ws.column_dimensions['A'].width = 30  # Key
            ws.column_dimensions['B'].width = 40  # 英文 Value
            ws.column_dimensions['C'].width = 40  # Value
            ws.column_dimensions['D'].width = 12  # 语言
            ws.column_dimensions['E'].width = 10  # 长度
            ws.column_dimensions['F'].width = 10  # 基准
            ws.column_dimensions['G'].width = 12  # 差异%
            
            # 设置数据行对齐和格式
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_num, cell in enumerate(row, 1):
                    if col_num in [4, 5, 6, 7]:  # 语言、长度、基准、差异列居中
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_num in [2, 3]:  # 英文 Value 和 Value 列左对齐，自动换行
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    elif col_num == 1:  # Key 列左对齐
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # 保存文件
            wb.save(file_path)
            
            # 更新按钮文字提示
            self.copy_btn.setText("已导出 ✓")
            from PyQt6.QtCore import QTimer
            QTimer.singleShot(2000, lambda: self.copy_btn.setText("导出 Excel"))
            
            # 显示成功提示
            from utils.toast import Toast
            Toast.show_toast(self, f"✅ 已导出到 {file_path}", 2000)
            
        except Exception as e:
            from utils.toast import Toast
            Toast.show_toast(self, f"❌ 导出失败: {str(e)}", 3000)
        
    def update_results(self, results: dict):
        """更新对比结果显示"""
        self.results = results
        
        if not results:
            # 无结果
            self.result_table.setVisible(False)
            self.empty_widget.setVisible(True)
            self.empty_label.setText("✓ 未发现变长的字段")
            self.empty_label.setStyleSheet("""
                font-size: 16px; 
                color: #34C759; 
                padding: 60px;
                font-weight: 500;
            """)
            self.stats_label.setText("所有字段长度都在合理范围内")
            self.stats_label.setStyleSheet("font-size: 14px; color: #34C759;")
            self.copy_btn.setVisible(False)
            return
        
        # 有结果
        self.empty_widget.setVisible(False)
        self.result_table.setVisible(True)
        self.result_table.setRowCount(0)
        
        # 按差异百分比排序（从大到小）
        self.sorted_results = sorted(
            results.items(),
            key=lambda x: x[1]['diff_percent'],
            reverse=True
        )
        
        # 更新统计
        total_count = len(results)
        max_diff = max(r['diff_percent'] for r in results.values())
        self.stats_label.setText(f"发现 {total_count} 个变长字段 · 最大增长 {max_diff:.0f}%")
        self.stats_label.setStyleSheet("font-size: 14px; color: #FF9500; font-weight: 500;")
        self.copy_btn.setVisible(True)
        self.copy_btn.setText("导出 Excel")  # 确保按钮文字正确
        
        # 填充表格
        self.result_table.setRowCount(len(self.sorted_results))
        
        for row, (_, data) in enumerate(self.sorted_results):
            # Key
            key_item = QTableWidgetItem(data['key'])
            self.result_table.setItem(row, 0, key_item)
            
            # 英文 Value（从 all_values 中获取）
            en_value = ""
            if 'all_values' in data:
                # 优先查找 'en'，如果没有则查找以 'en' 开头的语言代码
                if 'en' in data['all_values']:
                    en_value = data['all_values']['en']['value']
                else:
                    for lang_code in data['all_values'].keys():
                        if lang_code.startswith('en'):
                            en_value = data['all_values'][lang_code]['value']
                            break
            
            en_value_preview = en_value
            if len(en_value_preview) > 80:
                en_value_preview = en_value_preview[:80] + "..."
            en_value_item = QTableWidgetItem(en_value_preview)
            en_value_item.setToolTip(en_value)  # 完整内容在 tooltip 中
            self.result_table.setItem(row, 1, en_value_item)
            
            # Value（目标语言的 value）
            value_preview = data['target_value']
            if len(value_preview) > 80:
                value_preview = value_preview[:80] + "..."
            value_item = QTableWidgetItem(value_preview)
            value_item.setToolTip(data['target_value'])
            self.result_table.setItem(row, 2, value_item)
            
            # 语言
            lang_item = QTableWidgetItem(data['target_lang'])
            lang_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.result_table.setItem(row, 3, lang_item)
            
            # 长度
            len_item = QTableWidgetItem(str(data['target_length']))
            len_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.result_table.setItem(row, 4, len_item)
            
            # 基准
            base_len = data['base_length']
            if isinstance(base_len, float) and base_len.is_integer():
                base_str = str(int(base_len))
            else:
                base_str = f"{base_len:.1f}"
            base_item = QTableWidgetItem(base_str)
            base_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            base_item.setForeground(QBrush(QColor("#8E8E93")))
            self.result_table.setItem(row, 5, base_item)
            
            # 差异（根据严重程度着色）
            diff_percent = data['diff_percent']
            diff_text = f"+{diff_percent:.0f}%"
            diff_item = QTableWidgetItem(diff_text)
            diff_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            
            # 颜色梯度：50% 以下橙色，50% 以上红色
            if diff_percent >= 100:
                diff_item.setForeground(QBrush(QColor("#FF3B30")))
            elif diff_percent >= 50:
                diff_item.setForeground(QBrush(QColor("#FF6B35")))
            else:
                diff_item.setForeground(QBrush(QColor("#FF9500")))
            
            font = QFont()
            font.setBold(True)
            diff_item.setFont(font)
            self.result_table.setItem(row, 6, diff_item)
        
        # 行高
        for i in range(len(self.sorted_results)):
            self.result_table.setRowHeight(i, 44)
